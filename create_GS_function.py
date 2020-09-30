# import snowflake.connector as sfcon
# pip install -t lib google-auth google-auth-httplib2 google-api-python-client --upgrade
import Snowflake_Cred as cred
import glob
import os
import pandas as pd
import numpy as np
from google.oauth2.service_account import Credentials
import gspread
from gspread import utils
from gspread import models
import json
import datetime

from snowflake.sqlalchemy import URL
from sqlalchemy import create_engine
from sqlalchemy.sql import text
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


#Snowflake Credentials
my_user = cred.mySF["user"]
my_password = cred.mySF["password"]
my_account = cred.mySF["account"]
my_role = cred.mySF["role"] 
my_database = cred.mySF["database"]
my_schema = cred.mySF["schema"]
my_wh = cred.mySF["warehouse"]

#Google sheet authentication process
scopes = ['https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive']

token_file = 'token.json'


# Fill in your SFlake credential details 
engine = create_engine(URL(
    account = my_account,
    user = my_user,
    password = my_password,
    database = my_database,
    schema = my_schema,
    warehouse = my_wh,
    role=my_role,
))    

connection = engine.connect()

#Google Sheet Connection
credentials = Credentials.from_service_account_file(token_file, scopes=scopes)

gc = gspread.authorize(credentials)


def create_gs(txt_file,gs_sheet,ws_sheet): 
    f=open(txt_file,"r")
    query=f.read()

    df = pd.read_sql_query(text(query), engine) #Read query to Panda dataframe 
    df2=df.fillna(0) #Change null values to 0 
    
    col_header = [[y.replace("_"," ") for y in (x.upper() for x in df2.columns.values.tolist())]] #Capitalize column header
    
    #Get data range: Number of rows and colums and the column name in letter to load into Google Sheet
    rows = len(df2)+1
    cols = len(df2.columns)
    col_index = get_column_letter(cols)
    my_range = 'A1:{}{}'.format(col_index,rows)

    #Open Google Sheet
    my_ws = gc.open(gs_sheet).worksheet(ws_sheet) #The specific worksheet where data will be import

    #Clear all cell in sheet then update with new data
    my_ws.clear() 
    data_load = [df2.columns.values.tolist()] + df2.values.tolist()
    my_ws.update(my_range, data_load)  
    my_ws.update('A1:{}1'.format(col_index), col_header) #Update the header row to uppercase 
    print(my_range)

def query_df(query):     
    #This function create df from query insert. Ideal for short one line query.
    return pd.read_sql_query(text(query), engine) #Read query to Panda dataframe 

def txt_df(txt_file): 
    #This function create df from query saved in text file. Ideal for long query.
    f=open(txt_file,"r")
    query=f.read()
    return pd.read_sql_query(text(query), engine) #Read query to Panda dataframe  

def df_gs(df,gs_sheet,ws_sheet): 
    df2=df.fillna(0) #Change null values to 0 
    
    col_header = [[y.replace("_"," ") for y in (x.upper() for x in df2.columns.values.tolist())]] #Capitalize column header
    
    #Get data range: Number of rows and colums and the column name in letter to load into Google Sheet
    rows = len(df2)+1
    cols = len(df2.columns)
    col_index = get_column_letter(cols)
    my_range = 'A1:{}{}'.format(col_index,rows)

    #Open Google Sheet
    my_ws = gc.open(gs_sheet).worksheet(ws_sheet) #The specific worksheet where data will be import

    #Clear all cell in sheet then update with new data
    my_ws.clear() 
    data_load = [df2.columns.values.tolist()] + df2.values.tolist()
    my_ws.update(my_range, data_load)  
    my_ws.update('A1:{}1'.format(col_index), col_header) #Update the header row to uppercase 
    print(my_range)

def sheet_to_array(gs_sheet, ws_sheet):
    worksheet = gc.open(gs_sheet).worksheet(ws_sheet)
    return np.array(worksheet.get_all_values())    

def sheet_to_df(gs_sheet, ws_sheet):
    worksheet = gc.open(gs_sheet).worksheet(ws_sheet)
    return pd.DataFrame(worksheet.get_all_records())
